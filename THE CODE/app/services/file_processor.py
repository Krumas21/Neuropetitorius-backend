"""File processing service for various document types."""

import io
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from PIL import Image
from PyPDF2 import PdfReader


class FileProcessingService:
    """Service to extract text from various file types."""

    SUPPORTED_TYPES = {
        "pdf": "application/pdf",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xls": "application/vnd.ms-excel",
        "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "txt": "text/plain",
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "gif": "image/gif",
        "bmp": "image/bmp",
    }

    def extract_text(self, file_content: bytes, filename: str) -> str:
        """Extract text from a file based on its extension."""
        ext = Path(filename).suffix.lower().lstrip(".")

        if ext == "pdf":
            return self._extract_pdf(file_content)
        elif ext == "docx":
            return self._extract_docx(file_content)
        elif ext in ("xlsx", "xls"):
            return self._extract_excel(file_content)
        elif ext == "txt":
            return self._extract_txt(file_content)
        elif ext in ("png", "jpg", "jpeg", "gif", "bmp"):
            return self._extract_image(file_content)
        else:
            return f"[Unsupported file type: {ext}]"

    def _extract_pdf(self, content: bytes) -> str:
        """Extract text from PDF."""
        try:
            reader = PdfReader(io.BytesIO(content))
            text_parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
            return "\n\n".join(text_parts) if text_parts else "[Empty PDF]"
        except Exception as e:
            return f"[PDF extraction error: {str(e)}]"

    def _extract_docx(self, content: bytes) -> str:
        """Extract text from DOCX."""
        try:
            doc = Document(io.BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs) if paragraphs else "[Empty DOCX]"
        except Exception as e:
            return f"[DOCX extraction error: {str(e)}]"

    def _extract_excel(self, content: bytes) -> str:
        """Extract text from Excel."""
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            parts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                parts.append(f"=== Sheet: {sheet_name} ===")
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        parts.append(row_text)
            return "\n".join(parts) if parts else "[Empty Excel]"
        except Exception as e:
            return f"[Excel extraction error: {str(e)}]"

    def _extract_txt(self, content: bytes) -> str:
        """Extract text from plain text."""
        try:
            return content.decode("utf-8", errors="ignore")
        except Exception as e:
            return f"[Text extraction error: {str(e)}]"

    def _extract_image(self, content: bytes) -> str:
        """Extract basic info from image."""
        try:
            img = Image.open(io.BytesIO(content))
            info = [
                f"Image: {img.format}",
                f"Size: {img.size[0]}x{img.size[1]} pixels",
                f"Mode: {img.mode}",
            ]
            if hasattr(img, "info") and img.info:
                info.append(f"Info: {img.info}")
            return (
                "\n".join(info)
                + "\n[Note: Image content cannot be directly read as text. For full image analysis, consider using Gemini API vision capabilities.]"
            )
        except Exception as e:
            return f"[Image extraction error: {str(e)}]"


file_processor = FileProcessingService()
